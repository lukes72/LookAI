#!/usr/bin/env python3
"""Unified unattended runner for the paper and AI-news daily reports.

By default this command never sends to Feishu. A real send requires the
explicit ``--send`` flag and a passing local quality gate.
"""

from __future__ import annotations

import argparse
import json
import os
import subprocess
import sys
import urllib.error
import urllib.request
from datetime import date
from pathlib import Path
from typing import Any

import quality_review

try:
    import fitz  # PyMuPDF，用于读取论文 PDF 全文
except ImportError:  # pragma: no cover - 缺少依赖时回退到仅使用摘要
    fitz = None


BASE_DIR = Path(__file__).resolve().parent
REVIEW_JSON = BASE_DIR / "quality_review.json"
NEWS_JSON = BASE_DIR / "news_items.json"
PAPERS_JSON = BASE_DIR / "new_papers.json"
LLM_FILL_JSON = BASE_DIR / "llm_fill.json"


def run_script(script: str, *args: str) -> None:
    command = [sys.executable, str(BASE_DIR / script), *args]
    print(f"[RUN] {' '.join(command)}")
    subprocess.run(command, cwd=BASE_DIR, check=True)


def load_json(path: Path, default: Any) -> Any:
    return quality_review.load_json(path, default)


def request_llm(messages: list[dict[str, str]]) -> dict[str, Any]:
    api_base = os.environ.get("LLM_API_BASE", "https://api.openai.com/v1").rstrip("/")
    api_key = os.environ.get("LLM_API_KEY", "").strip()
    model = os.environ.get("LLM_MODEL", "gpt-4o-mini").strip()
    if not api_key:
        raise RuntimeError("未配置 LLM_API_KEY，无法自动生成中文总结。")

    payload = json.dumps(
        {"model": model, "temperature": 0.1, "response_format": {"type": "json_object"}, "messages": messages},
        ensure_ascii=False,
    ).encode("utf-8")
    request = urllib.request.Request(
        f"{api_base}/chat/completions",
        data=payload,
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            body = json.loads(response.read().decode("utf-8"))
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as exc:
        raise RuntimeError(f"LLM 请求失败：{exc}") from exc

    content = body.get("choices", [{}])[0].get("message", {}).get("content", "")
    if isinstance(content, list):
        content = "".join(part.get("text", "") for part in content if isinstance(part, dict))
    try:
        result = json.loads(str(content))
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"LLM 返回不是有效 JSON：{content[:200]}") from exc
    if not isinstance(result, dict):
        raise RuntimeError("LLM 返回 JSON 不是对象。")
    return result


def extract_pdf_text(pdf_path: str, max_chars: int = 8000) -> str:
    """读取论文 PDF 正文，返回截断后的纯文本；失败时返回空字符串。"""
    if not fitz or not pdf_path:
        return ""
    pdf = Path(pdf_path)
    if not pdf.exists():
        return ""
    try:
        doc = fitz.open(str(pdf))
        parts: list[str] = []
        total = 0
        for page in doc:
            text = (page.get_text("text") or "").strip()
            if not text:
                continue
            if total + len(text) > max_chars:
                parts.append(text[: max_chars - total])
                total = max_chars
                break
            parts.append(text)
            total += len(text)
        doc.close()
        return "\n".join(parts).strip()
    except Exception:
        return ""


def summarize_papers() -> None:
    state = load_json(PAPERS_JSON, {})
    papers = state.get("papers_to_process", []) if isinstance(state, dict) else []
    if not papers:
        return

    results: dict[str, dict[str, str]] = {}
    for paper in papers:
        arxiv_id = str(paper.get("arxiv_id", "")).strip()
        full_text = extract_pdf_text(paper.get("pdf_local_path", ""))
        prompt = {
            "role": "user",
            "content": json.dumps(
                {
                    "task": (
                        "请基于论文标题、摘要和正文，输出严格 JSON。"
                        "summary_cn 用中文结构化总结，逐行给出以下字段（材料缺失时写“未提供”）：\n"
                        "一句话：一句话概括论文核心贡献\n"
                        "动机：论文要解决的问题\n"
                        "方法：提出的方法或技术\n"
                        "结果：关键实验结果或发现\n"
                        "结论：论文的最终结论与意义\n"
                        "affiliations 从作者信息或正文首页提取作者所属单位，用分号分隔；"
                        "无法确定则返回空字符串。只使用给定材料，不得臆造数字、单位或性能数据。"
                    ),
                    "paper": {
                        "arxiv_id": arxiv_id,
                        "title": paper.get("title", ""),
                        "authors": paper.get("authors", ""),
                        "abstract": paper.get("abstract") or paper.get("summary", ""),
                        "full_text": full_text,
                    },
                    "output": {"summary_cn": "", "affiliations": ""},
                },
                ensure_ascii=False,
            ),
        }
        result = request_llm(
            [
                {"role": "system", "content": "你是严谨的 AI 论文编辑，精通机器学习与自然语言处理。只使用输入材料中的事实，不做未经证实的推断。"},
                prompt,
            ]
        )
        results[arxiv_id] = {
            "summary_cn": str(result.get("summary_cn", "")).strip(),
            "affiliations": str(result.get("affiliations", "")).strip(),
        }

    LLM_FILL_JSON.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    run_script("fill_llm.py", "--json", str(LLM_FILL_JSON))
    run_script("monitor.py", "--sync-pending-state")


def summarize_news() -> None:
    state = load_json(NEWS_JSON, {})
    items = state.get("items", []) if isinstance(state, dict) else []
    if not items:
        return

    changed = False
    for item in items:
        if str(item.get("summary_cn", "")).strip():
            continue
        result = request_llm(
            [
                {"role": "system", "content": "你是严谨的 AI 新闻编辑。只能根据标题、来源和原文摘要写 1 至 2 句中文，不补充未经提供的事实。"},
                {
                    "role": "user",
                    "content": json.dumps(
                        {
                            "title": item.get("title", ""),
                            "source": item.get("source", ""),
                            "url": item.get("url", ""),
                            "source_summary": item.get("summary_short", ""),
                            "output": {"summary_cn": ""},
                        },
                        ensure_ascii=False,
                    ),
                },
            ]
        )
        item["summary_cn"] = str(result.get("summary_cn", "")).strip()
        changed = True

    if changed:
        state["items"] = items
        NEWS_JSON.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def load_papers_for_review() -> list[dict[str, Any]]:
    try:
        import openpyxl
    except ImportError:
        return []
    excel = BASE_DIR / "papers_record.xlsx"
    if not excel.exists():
        return []
    workbook = openpyxl.load_workbook(excel, read_only=True)
    if "Papers" not in workbook.sheetnames:
        return []
    sheet = workbook["Papers"]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "") for value in rows[0]]
    return [
        {header: (row[index] if index < len(row) else "") for index, header in enumerate(headers)}
        for row in rows[1:]
        if row and row[0]
    ]


def review_current_state() -> dict[str, Any]:
    papers_report = quality_review.review_papers(load_papers_for_review())
    news_state = load_json(NEWS_JSON, {})
    news_items = news_state.get("items", []) if isinstance(news_state, dict) else []
    news_report = quality_review.review_news(news_items, news_state.get("errors", []) if isinstance(news_state, dict) else [])
    report = quality_review.merge_reports(papers_report, news_report)
    quality_review.write_review_report(report, REVIEW_JSON)
    print(f"[QUALITY] ok={report['ok']} errors={len(report['errors'])} warnings={len(report['warnings'])}")
    for error in report["errors"]:
        print(f"[QUALITY][ERROR] {error}")
    for warning in report["warnings"]:
        print(f"[QUALITY][WARN] {warning}")
    return report


def send_reports(target_date: str, papers_only: bool, news_only: bool) -> None:
    if not news_only:
        run_script("send_feishu.py", "--date", target_date)
    if not papers_only:
        run_script("send_news_feishu.py")


def main() -> int:
    parser = argparse.ArgumentParser(description="统一运行 AI 论文与新闻日报")
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--dry-run", action="store_true", help="运行检查并打印结果，不发送")
    mode.add_argument("--no-send", action="store_true", help="运行抓取和检查，不发送")
    mode.add_argument("--send", action="store_true", help="质量检查通过后发送到飞书")
    parser.add_argument("--date", default=date.today().isoformat(), help="日报日期 YYYY-MM-DD")
    parser.add_argument("--hours", type=int, default=24, help="新闻时间窗口")
    parser.add_argument("--skip-fetch", action="store_true", help="跳过网络抓取，使用现有本地数据")
    parser.add_argument("--papers-only", action="store_true")
    parser.add_argument("--news-only", action="store_true")
    args = parser.parse_args()

    if args.papers_only and args.news_only:
        parser.error("--papers-only 与 --news-only 不能同时使用")
    if not args.send:
        print("[SAFE] 未指定 --send，本次不会调用飞书发送 API。")

    if not args.skip_fetch:
        if not args.news_only:
            run_script("monitor.py")
        if not args.papers_only:
            run_script("news_monitor.py", "--hours", str(args.hours))

    try:
        if not args.news_only:
            summarize_papers()
        if not args.papers_only:
            summarize_news()
    except RuntimeError as exc:
        print(f"[LLM][ERROR] {exc}")
        print("[SAFE] LLM 未完成，禁止发送。")
        return 2

    if not args.news_only:
        run_script("viewer/build_data.py")
    report = review_current_state()
    if args.send:
        if not report["ok"]:
            print("[SAFE] 质量门禁未通过，禁止发送。")
            return 3
        send_reports(args.date, args.papers_only, args.news_only)
    return 0 if report["ok"] else (1 if args.dry_run or args.no_send or args.skip_fetch else 1)


if __name__ == "__main__":
    raise SystemExit(main())

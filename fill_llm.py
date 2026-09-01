#!/usr/bin/env python3
"""把 LLM 生成的 summary_cn / affiliations 回填到 papers_record.xlsx。

用法:
  python fill_llm.py --json llm_fill.json

llm_fill.json 格式:
  {
    "2608.28113": {"affiliations": "MIT; Stanford", "summary_cn": "本文提出..."},
    ...
  }
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent
EXCEL_FILE = BASE_DIR / "papers_record.xlsx"


def main() -> None:
    parser = argparse.ArgumentParser(description="回填 LLM 补全结果到 Excel")
    parser.add_argument("--json", default="llm_fill.json", help="LLM 补全结果文件路径")
    args = parser.parse_args()

    fill_path = Path(args.json)
    if not fill_path.is_absolute():
        fill_path = BASE_DIR / fill_path
    if not fill_path.exists():
        print(f"[ERROR] 找不到 {fill_path}")
        sys.exit(1)

    data = json.loads(fill_path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        print("[ERROR] llm_fill.json 必须是 {arxiv_id: {...}} 字典")
        sys.exit(1)

    from openpyxl import load_workbook

    if not EXCEL_FILE.exists():
        print("[ERROR] 找不到 papers_record.xlsx")
        sys.exit(1)

    wb = load_workbook(EXCEL_FILE)
    if "Papers" not in wb.sheetnames:
        print("[ERROR] Excel 缺少 Papers 工作表")
        sys.exit(1)
    ws = wb["Papers"]

    headers = [str(c.value) if c.value is not None else "" for c in ws[1]]
    header_index = {name: i + 1 for i, name in enumerate(headers)}
    if "arxiv_id" not in header_index:
        print("[ERROR] Excel 缺少 arxiv_id 列")
        sys.exit(1)

    updated = 0
    for arxiv_id, fields in data.items():
        row_num = None
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=header_index["arxiv_id"]).value
            if val is not None and str(val).strip() == str(arxiv_id).strip():
                row_num = r
                break
        if row_num is None:
            print(f"[WARN] Excel 中未找到 {arxiv_id}")
            continue
        for key in ("affiliations", "summary_cn"):
            if key in fields and key in header_index:
                ws.cell(row=row_num, column=header_index[key], value=str(fields[key]).strip())
        updated += 1
        print(f"[OK] 已回填 {arxiv_id}")

    wb.save(EXCEL_FILE)
    print(f"[DONE] 共回填 {updated} 篇论文 -> {EXCEL_FILE}")


if __name__ == "__main__":
    main()

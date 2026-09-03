#!/usr/bin/env python3
"""飞书论文日报发送脚本（机器人模式）。

从 papers_record.xlsx 读取指定日期抓取的论文，生成飞书日报，
并直接调用飞书开放平台 API 以应用机器人身份发送给指定用户。

用法:
  python send_feishu.py                 # 发送今天的论文日报
  python send_feishu.py --dry-run       # 只预览消息，不真正发送
  python send_feishu.py --date 2026-08-31
  python send_feishu.py --all           # 发送全部论文（不限于当天）

配置（优先级从高到低）:
  环境变量 FEISHU_APP_ID / FEISHU_APP_SECRET / FEISHU_USER_ID
  或仓库根目录 feishu_config.json:
    {"app_id": "cli_xxx", "app_secret": "xxx", "user_id": "ou_xxx"}
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import date
from pathlib import Path

import requests

from monitor import is_relevant as monitor_is_relevant

BASE_DIR = Path(__file__).resolve().parent
EXCEL_FILE = BASE_DIR / "papers_record.xlsx"
CONFIG_FILE = BASE_DIR / "feishu_config.json"

FEISHU_BASE = "https://open.feishu.cn/open-apis"

EXCEL_HEADERS = [
    "arxiv_id",
    "title",
    "authors",
    "affiliations",
    "published_date",
    "categories",
    "abstract",
    "summary_cn",
    "pdf_filename",
    "crawled_date",
    "notes",
]

# arXiv 分类 -> 中文标签（最多取前 3 个，其余忽略）
CATEGORY_LABELS = {
    "cs.AI": "人工智能",
    "cs.CL": "自然语言处理",
    "cs.LG": "机器学习",
    "cs.CV": "计算机视觉",
    "cs.DB": "数据库",
    "cs.IR": "信息检索",
    "cs.MA": "多智能体",
    "cs.SE": "软件工程",
    "cs.RO": "机器人",
    "cs.NE": "神经网络",
    "cs.SY": "系统与控制",
    "cs.CR": "安全",
    "cs.DS": "数据科学",
}

# 结构化总结标签（多行总结渲染时保留中文标签，不添加 emoji）
SUMMARY_LABELS = (
    "一句话",
    "问题",
    "动机",
    "方法",
    "结果",
    "亮点",
    "贡献",
    "结论",
)

SEPARATOR = "─" * 20


def load_config() -> dict:
    cfg: dict = {}
    if CONFIG_FILE.exists():
        try:
            cfg = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
        except Exception as e:  # noqa: BLE001
            print(f"[WARN] 读取 feishu_config.json 失败: {e}")
    return cfg


def resolve_setting(env_name: str, cfg_key: str, cfg: dict) -> str:
    return str(os.environ.get(env_name) or cfg.get(cfg_key, "")).strip()


def load_papers() -> list[dict]:
    """从 Excel 读取全部论文（含空值清洗）。"""
    if not EXCEL_FILE.exists():
        return []

    from openpyxl import load_workbook

    wb = load_workbook(EXCEL_FILE, read_only=True)
    if "Papers" not in wb.sheetnames:
        return []
    ws = wb["Papers"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h) if h is not None else "" for h in rows[0]]
    idx = {name: i for i, name in enumerate(headers)}

    def raw(row, name: str) -> str:
        i = idx.get(name)
        if i is None or i >= len(row):
            return ""
        v = row[i]
        return "" if v is None else str(v).strip()

    def cell(row, name: str) -> str:
        # 普通字段：压成单行，避免换行打乱排版
        return raw(row, name).replace("\n", " ").strip()

    papers: list[dict] = []
    for row in rows[1:]:
        p = {h: cell(row, h) for h in EXCEL_HEADERS}
        # summary_cn / abstract 保留换行，便于渲染结构化多行总结
        p["summary_cn"] = raw(row, "summary_cn")
        p["abstract"] = raw(row, "abstract")
        papers.append(p)
    return papers


def paper_is_relevant(p: dict) -> bool:
    """复用 monitor.py 的领域黑名单，过滤临床/医疗/生物/农业/金融/审计等应用方向。"""
    return monitor_is_relevant({
        "title": p.get("title", ""),
        "summary": p.get("abstract", ""),
        "categories": p.get("categories", ""),
    })


def category_summary(categories: str, limit: int = 3) -> str:
    """把 arXiv 分类转成短中文标签，如 `cs.AI, cs.CL` -> `人工智能 · 自然语言处理`。"""
    if not categories:
        return ""
    parts = [c.strip() for c in categories.replace(";", ",").split(",") if c.strip()]
    labels: list[str] = []
    for part in parts:
        label = CATEGORY_LABELS.get(part, part)
        if label not in labels:
            labels.append(label)
        if len(labels) >= limit:
            break
    return " · ".join(labels)


def author_summary(authors: str, limit: int = 3) -> str:
    """作者只展示前 3 位，后面用 `等 N 人` 概括，让标题区更快可读。"""
    if not authors:
        return ""
    names = [n.strip() for n in authors.replace(";", ",").split(",") if n.strip()]
    if len(names) <= limit:
        return ", ".join(names)
    return ", ".join(names[:limit]) + f" 等 {len(names)} 人"


def summary_lines(summary_cn: str) -> list[str]:
    """把中文总结渲染成无 emoji 的易读行，保留中文标签与原文内容。"""
    if not summary_cn:
        return []

    lines = [ln.strip() for ln in summary_cn.splitlines() if ln.strip()]
    if not lines:
        return []

    label_pattern = re.compile(
        r"^(" + "|".join(re.escape(t) for t in SUMMARY_LABELS) + r")[:：]\s*(.+)$"
    )

    rendered: list[str] = []
    for line in lines:
        m = label_pattern.match(line)
        if m:
            label, text = m.group(1), m.group(2)
            rendered.append(f"{label}：{text}")
        else:
            rendered.append(line)

    return rendered


def paper_lines(papers: list[dict]) -> list[str]:
    """生成日报正文各行（不含头部标题行）。"""
    lines: list[str] = []
    for i, p in enumerate(papers, start=1):
        title = p.get("title") or "(无标题)"
        lines.append(f"【{i}】{title}")

        if p.get("affiliations"):
            lines.append(f"机构：{p['affiliations']}")

        author_text = author_summary(p.get("authors", ""))
        if author_text:
            lines.append(f"作者：{author_text}")

        meta_parts = []
        if p.get("published_date"):
            meta_parts.append(p["published_date"])
        if p.get("arxiv_id"):
            meta_parts.append(f"arXiv {p['arxiv_id']}")
        cat = category_summary(p.get("categories", ""))
        if cat:
            meta_parts.append(cat)
        if meta_parts:
            lines.append(" · ".join(meta_parts))

        if p.get("arxiv_id"):
            lines.append(f"PDF: https://arxiv.org/pdf/{p['arxiv_id']}")

        lines.extend(summary_lines(p.get("summary_cn", "")))

        if i < len(papers):
            lines.append(SEPARATOR)

    return lines


def build_markdown(papers: list[dict], target_date: str) -> str:
    lines = [
        f"论文日报 | {target_date}",
        f"共 {len(papers)} 篇 · 已按你关注的方向筛选",
        "",
    ] + paper_lines(papers)
    return "\n".join(lines).strip()


def line_to_elements(line: str) -> list[dict]:
    """把一行文本转换为飞书 post 富文本元素。"""
    if line.startswith("PDF: "):
        url = line[len("PDF: "):].strip()
        return [
            {"tag": "text", "text": "PDF: "},
            {"tag": "a", "text": url, "href": url},
        ]
    return [{"tag": "text", "text": line}]


def build_post_content(papers: list[dict], target_date: str) -> dict:
    content: list[list[dict]] = []

    def add_line(line: str) -> None:
        elements = line_to_elements(line)
        if elements:
            content.append(elements)

    add_line(f"论文日报 | {target_date}")
    add_line(f"共 {len(papers)} 篇 · 已按你关注的方向筛选")
    add_line("")

    for line in paper_lines(papers):
        add_line(line)

    return {
        "zh_cn": {
            "title": f"论文日报 | {target_date}",
            "content": content,
        }
    }


def get_tenant_access_token(app_id: str, app_secret: str) -> str:
    url = f"{FEISHU_BASE}/auth/v3/tenant_access_token/internal"
    resp = requests.post(
        url, json={"app_id": app_id, "app_secret": app_secret}, timeout=30
    )
    resp.raise_for_status()
    data = resp.json()
    if data.get("code") != 0:
        raise RuntimeError(f"获取 tenant_access_token 失败: {data}")
    return data["tenant_access_token"]


def send_post_message(token: str, user_id: str, post_content: dict) -> dict:
    url = f"{FEISHU_BASE}/im/v1/messages?receive_id_type=open_id"
    payload = {
        "receive_id": user_id,
        "msg_type": "post",
        "content": json.dumps(post_content, ensure_ascii=False),
    }
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json; charset=utf-8",
    }
    resp = requests.post(url, headers=headers, json=payload, timeout=30)
    resp.raise_for_status()
    data = resp.json()
    if data.get("code") != 0:
        raise RuntimeError(f"发送消息失败: {data}")
    return data.get("data", {})


def send_feishu(
    markdown: str,
    post_content: dict,
    user_id: str,
    app_id: str,
    app_secret: str,
    dry_run: bool = False,
) -> bool:
    if dry_run:
        print("[DRY-RUN] 将调用飞书开放平台 API 发送 post 消息")
        print(f"[DRY-RUN] user_id={user_id}, 内容长度={len(markdown)}")
        print("----- 消息内容预览 -----")
        print(markdown)
        print("------------------------")
        return True

    print("[INFO] 正在获取 tenant_access_token ...")
    token = get_tenant_access_token(app_id, app_secret)
    print(f"[INFO] 正在发送 {len(markdown)} 字符的日报 ...")
    result = send_post_message(token, user_id, post_content)
    print(f"[OK] 已发送到飞书: message_id={result.get('message_id', '')}")
    return True


def main() -> None:
    parser = argparse.ArgumentParser(description="发送飞书论文日报")
    parser.add_argument("--dry-run", action="store_true", help="只预览，不发送")
    parser.add_argument("--date", default=date.today().isoformat(), help="指定日期 YYYY-MM-DD")
    parser.add_argument("--all", action="store_true", help="发送全部论文而不是仅当天")
    args = parser.parse_args()

    cfg = load_config()
    app_id = resolve_setting("FEISHU_APP_ID", "app_id", cfg)
    app_secret = resolve_setting("FEISHU_APP_SECRET", "app_secret", cfg)
    user_id = resolve_setting("FEISHU_USER_ID", "user_id", cfg)

    if not app_id or not app_secret:
        print("[ERROR] 未配置 FEISHU_APP_ID / FEISHU_APP_SECRET（或 feishu_config.json 中的 app_id/app_secret）")
        sys.exit(1)
    if not user_id:
        print("[ERROR] 未配置 FEISHU_USER_ID 或 feishu_config.json 中的 user_id")
        sys.exit(1)

    papers = load_papers()
    if args.all:
        selected = papers
    else:
        selected = [p for p in papers if p.get("crawled_date") == args.date]

    before = len(selected)
    selected = [p for p in selected if paper_is_relevant(p)]
    if len(selected) != before:
        print(f"[FILTER] 已过滤 {before - len(selected)} 篇非目标方向论文（临床/医疗/生物/农业/金融/审计等）")

    if not selected:
        markdown = f"今日（{args.date}）未发现新的 AI 论文。"
        post_content = {
            "zh_cn": {
                "title": f"论文日报 | {args.date}",
                "content": [[{"tag": "text", "text": markdown}]],
            }
        }
    else:
        markdown = build_markdown(selected, args.date)
        post_content = build_post_content(selected, args.date)

    ok = send_feishu(markdown, post_content, user_id, app_id, app_secret, dry_run=args.dry_run)
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
